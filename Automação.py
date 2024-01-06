from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from decimal import Decimal
import openpyxl
import pyperclip


def ainda_existem_produtos_na_planilha(linha_atual, planilha):
    # Verificar se a próxima linha da planilha não é vazia
    proxima_linha = linha_atual + 1
    return planilha.cell(row=proxima_linha, column=1).value is not None

workbook = openpyxl.load_workbook ('Produtos.xlsx')
sheet_produtos = workbook["Página1"]


chromedriver_path = "C:\\Users\\jgabr\OneDrive\\Documentos\\chromedriver.exe"

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument(f"webdriver.chrome.driver={chromedriver_path}")

browser = webdriver.Chrome(options=chrome_options)


browser.get('https://joaogabri3ls.github.io/Formulario/')
# assert 'Google' in browser.title

for linha in sheet_produtos.iter_rows(min_row=2):

    #Coluna do nome
    nome = linha[0].value
    pyperclip.copy(nome)
    elem = browser.find_element(By.NAME, 'n') 
    elem.send_keys(Keys.CONTROL + 'v')

    #Coluna da descrição
    descricao = linha[1].value
    pyperclip.copy(descricao)
    elem = browser.find_element(By.NAME, 'd') 
    elem.send_keys(Keys.CONTROL + 'v')


    #Coluna da quantidade
    quantidade = int(linha[2].value)
    elem_quantidade = browser.find_element(By.NAME, 'quantidade')
    select_quantidade = Select(elem_quantidade)
    select_quantidade.select_by_value(str(quantidade))

    preco = linha[3].value
    preco = preco.replace('R$', '').replace(',', '.').strip()
    preco_decimal = Decimal(preco)
    pyperclip.copy(str(preco_decimal))
    elem = browser.find_element(By.NAME, 'p') 
    elem.send_keys(Keys.CONTROL + 'v')

    botao_enviar = browser.find_element(By.NAME, 'enviar') 
    WebDriverWait(browser, 10).until(EC.invisibility_of_element_located((By.ID, 'overlay')))
    botao_enviar.click()

    try:
        mensagem_confirmacao = WebDriverWait(browser, 10).until(
            EC.presence_of_element_located((By.ID, 'message-box'))
        )



        # Verificar se existem mais produtos na planilha
        if ainda_existem_produtos_na_planilha(linha[0].row, sheet_produtos):
            botao_sim = browser.find_element(By.NAME, 'sim')
            botao_sim.click()

        else:
            botao_nao = browser.find_element(By.NAME, 'nao')
            botao_nao.click()


    except Exception as e:
        print(f"Erro ao lidar com a mensagem de confirmação: {e}")

